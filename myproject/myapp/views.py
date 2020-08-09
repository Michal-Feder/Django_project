import calendar
from datetime import datetime

from django.shortcuts import render

from myapp import db
import json
from django.http import HttpResponse, HttpResponseRedirect
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, NamedStyle

from myapp.forms import DateForm


def index(request):
    print('in Index')
    return HttpResponse("Hello, world. You're at the myapp index.")

def EmployersListToJson(request):
    employers = db.selectEmployerJson()
    dump = json.dumps(employers)
    return HttpResponse(dump, content_type='application/json')

def EmployersListToExcel(request):
    wb = Workbook()
    ws = wb.active
    employers = db.selectEmployer()
    ws['A1'] = "Id"
    ws['B1'] = "Name"
    ws['C1'] = "Phone"
    ws['D1'] = "Email"

    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.border = Border(bottom=Side(border_style="thin"))
    header.alignment = Alignment(horizontal="center", vertical="center")

    header_row = ws[1]
    for cell in header_row:
        cell.style = header

    row = 1
    for employ in employers:
        ws['A'+str(row)]
        row += 1
        ws.append([employ.id,employ.name,employ.phone,employ.email])
    wb.save("Employers.xlsx")
    return HttpResponse("succeeded to write an excel sheet, called - 'Employers.xlsx'")

def day_of_date(request):
    if request.method == 'POST':
        form = DateForm(request.POST)
        result = calendar.monthrange(int(form.data['date'].split('-')[0]), int(form.data['date'].split('-')[1]))[1]
        month = form.data['date'].split('-')[1]
        return render(request,'result.html',{'result':result,'month':month})
    else:
        form=DateForm()
        return render(request,'form.html',{'form':form})

